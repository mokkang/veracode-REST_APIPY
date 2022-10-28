#!/usr/bin/python3
import requests
import openpyxl
import time

#CONSTANTS
API_URL = 'https://api.veracode.com/api/authn/v2/'


def extract_json(search_term):

    parameters = {
        'fullText': 'true',
        'fields': 'Account;Login Name;User Name;Email Address;Custom ID;Title;Login Enabled;Created;Is Active;Is saml User;Is Platform User;Has Elearning Role;Last Login;Roles;Teams'
    }

    response = requests.get(API_URL + search_term, params=parameters)

    json = response.json()

    if response.status_code != 200:
        return 'error'

    return json[0]


def write_account():
    filename = 'test_account.xlsx'

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    wc = ws['A']

    i = 1
    for cell in wc[1:]:
        i += 1
        account_details = extract_json(cell.value)
        time.sleep(10)
        ws['B' + str(i)] = account_details['Account']
        ws['C' + str(i)] = account_details['Login Name']
        ws['D' + str(i)] = account_details['User Name']

    wb.save(filename)


write_account()
